#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import io
import os
import uuid
from typing import Optional

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = FastAPI(title="排班工时计算器", version="1.0.0")

cors_origins = [
    o.strip()
    for o in os.getenv(
        "CORS_ORIGINS",
        "https://schedule-app-frontend-lilac.vercel.app,http://localhost:5173",
    ).split(",")
    if o.strip()
]
cors_origin_regex = os.getenv("CORS_ORIGIN_REGEX", r"https://.*\.vercel\.app")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_origin_regex=cors_origin_regex,
    allow_methods=["*"],
    allow_headers=["*"],
)

_file_cache: dict[str, bytes] = {}

# ─────────────────────────────────────────────
# 班次 / 颜色 工具
# ─────────────────────────────────────────────

SHIFT_TEXT_MAP = {
    "值": "值班", "值班": "值班",
    "白": "白班", "白班": "白班",
    "休": "休息", "休息": "休息",
    "年": "年假", "年假": "年假",
}

SHIFT_HOURS = {"值班": 12, "白班": 8, "休息": 0, "年假": 8}

SHIFT_FILL = {
    "值班": PatternFill(fill_type="solid", fgColor="FF0000"),
    "白班": PatternFill(fill_type="solid", fgColor="FFFF00"),
    "休息": PatternFill(fill_type="solid", fgColor="70AD47"),
    "年假": PatternFill(fill_type="solid", fgColor="FFC000"),
}

SUMMARY_KEYWORDS = {
    "应出勤": "should_work",
    "实际出勤": "actual_work",
    "本月余": "month_balance",
    "累计余": "total_balance",
    "应勤": "should_work",
    "实勤": "actual_work",
    "月余": "month_balance",
    "累余": "total_balance",
}

# 星期文字 -> 是否周末
WEEKEND_WORDS = {"六", "日", "六", "天", "Sat", "Sun", "Saturday", "Sunday", "6", "7"}
WEEKDAY_MAP = {
    "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7,
    "Mon": 1, "Tue": 2, "Wed": 3, "Thu": 4, "Fri": 5, "Sat": 6, "Sun": 7,
}


def get_bg_argb(cell):
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        c = fill.fgColor
        if c and c.type == "rgb":
            return c.rgb.upper()
    return None


def classify_by_text(cell):
    val = str(cell.value).strip() if cell.value is not None else ""
    return SHIFT_TEXT_MAP.get(val)


def classify_by_color(cell):
    argb = get_bg_argb(cell)
    if not argb:
        return None
    rgb = argb[-6:]
    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    if r >= 240 and g >= 240 and b >= 240:
        return None
    if r >= 180 and g <= 120 and b <= 120:
        return "值班"
    if r >= 180 and g >= 150 and b <= 120:
        return "白班"
    if b >= 150 and r <= 130 and g <= 160:
        return "休息"
    if g >= 130 and r <= 150 and b <= 120:
        return "休息"  # 绿色也算休息
    return None


def find_name_row(sheet, name):
    for row in sheet.iter_rows():
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if val == name:
                return cell.row, cell.column
    return None, None


def find_date_and_weekday_rows(sheet, name_row: int):
    """
    在 name_row 之前找：
      - date_row: 内容为 1-31 数字的行 -> {col: day_str}
      - weekday_row: 内容为 一二三...日 的行 -> {col: weekday_num(1-7)}
    """
    date_row_map = {}    # col -> "1".."31"
    weekday_map = {}     # col -> 1..7 (1=Mon, 7=Sun)

    date_best, date_score = None, 0
    week_best, week_score = None, 0

    for ri in range(1, name_row + 1):
        cells = list(sheet.iter_rows(min_row=ri, max_row=ri, values_only=False))[0]
        d_score = w_score = 0
        for c in cells:
            if c.value is None:
                continue
            val = str(c.value).strip()
            # 日期行判断
            try:
                n = int(val)
                if 1 <= n <= 31:
                    d_score += 1
            except (ValueError, TypeError):
                pass
            # 星期行判断
            for kw in WEEKDAY_MAP:
                if kw in val:
                    w_score += 1
                    break
        if d_score > date_score:
            date_score, date_best = d_score, ri
        if w_score > week_score:
            week_score, week_best = w_score, ri

    # 填充 date_row_map
    if date_best:
        for cell in list(sheet.iter_rows(min_row=date_best, max_row=date_best, values_only=False))[0]:
            if cell.value is not None:
                try:
                    n = int(str(cell.value).strip())
                    if 1 <= n <= 31:
                        date_row_map[cell.column] = str(n)
                except (ValueError, TypeError):
                    pass

    # 填充 weekday_map
    if week_best and week_best != date_best:
        for cell in list(sheet.iter_rows(min_row=week_best, max_row=week_best, values_only=False))[0]:
            if cell.value is not None:
                val = str(cell.value).strip()
                for kw, num in WEEKDAY_MAP.items():
                    if kw in val:
                        weekday_map[cell.column] = num
                        break

    return date_row_map, weekday_map


def find_summary_cols(sheet, header_row: int, name_col: int) -> dict:
    result = {}
    for ri in range(1, min(header_row + 3, sheet.max_row + 1)):
        row_cells = list(sheet.iter_rows(min_row=ri, max_row=ri, values_only=False))[0]
        found = {}
        for cell in row_cells:
            val = str(cell.value).strip() if cell.value else ""
            for kw, key in SUMMARY_KEYWORDS.items():
                if kw in val:
                    found[key] = cell.column
                    break
        if len(found) >= 2:
            result = found
            break
    return result


def read_summary_values(sheet, data_row: int, summary_cols: dict) -> dict:
    values = {
        "should_work": None,
        "actual_work": None,
        "month_balance": None,
        "total_balance": None,
    }
    for key, col in summary_cols.items():
        cell = sheet.cell(row=data_row, column=col)
        if cell.value is not None:
            try:
                values[key] = float(cell.value)
            except (ValueError, TypeError):
                values[key] = None
    return values


def calc_sheet(sheet, name: str, holidays: list[str] = None) -> Optional[dict]:
    """
    holidays: 节假日列表，格式 ["1", "2", ...] 表示该月第几天是节假日
    """
    if holidays is None:
        holidays = []
    holiday_set = set(str(h) for h in holidays)

    row_idx, name_col = find_name_row(sheet, name)
    if row_idx is None:
        return None

    # 找日期行和星期行
    date_row_map, weekday_map = find_date_and_weekday_rows(sheet, row_idx)

    # 找汇总栏位
    summary_cols = find_summary_cols(sheet, row_idx, name_col)
    summary_min_col = min(summary_cols.values()) if summary_cols else sheet.max_column + 1
    min_data_col = name_col + 1
    max_data_col = summary_min_col - 1 if summary_cols else sheet.max_column

    # 检测识别模式
    text_hits = color_hits = 0
    for row in sheet.iter_rows(min_row=row_idx, max_row=row_idx,
                               min_col=min_data_col, max_col=max_data_col):
        for cell in row:
            if classify_by_text(cell):
                text_hits += 1
            elif classify_by_color(cell):
                color_hits += 1
    use_text = text_hits >= color_hits

    counts = {"值班": 0, "白班": 0, "休息": 0, "周末休息": 0, "节假日休息": 0, "年假": 0}
    daily = []

    for row in sheet.iter_rows(min_row=row_idx, max_row=row_idx,
                               min_col=min_data_col, max_col=max_data_col):
        for cell in row:
            shift = classify_by_text(cell) if use_text else classify_by_color(cell)
            if shift is None:
                shift = classify_by_color(cell) if use_text else classify_by_text(cell)

            if shift not in ("值班", "白班", "休息", "年假"):
                continue

            col = cell.column
            day_str = date_row_map.get(col, str(col - name_col))
            weekday = weekday_map.get(col)  # 1-7, 6=Sat, 7=Sun
            is_weekend = weekday in (6, 7) if weekday else False
            is_holiday = day_str in holiday_set

            # 细分休息类型
            rest_type = None
            if shift == "休息":
                if is_holiday:
                    rest_type = "节假日休息"
                elif is_weekend:
                    rest_type = "周末休息"
                else:
                    rest_type = "休息"
                counts[rest_type] += 1
            else:
                counts[shift] += 1

            daily.append({
                "col": col,
                "day": day_str,
                "shift": shift,
                "rest_type": rest_type,   # 仅休息时有值
                "hours": SHIFT_HOURS[shift],
                "is_weekend": is_weekend,
                "is_holiday": is_holiday,
                "weekday": weekday,
            })

    actual_work = counts["值班"] * 12 + counts["白班"] * 8 + counts["年假"] * 8
    doc_values = read_summary_values(sheet, row_idx, summary_cols)
    should_work = doc_values["should_work"]
    month_balance = doc_values["month_balance"]
    total_balance = doc_values["total_balance"]
    if month_balance is None and should_work is not None:
        month_balance = actual_work - should_work

    return {
        "actual_work": actual_work,
        "should_work": should_work,
        "month_balance": month_balance,
        "total_balance": total_balance,
        "duty": counts["值班"],
        "day_shift": counts["白班"],
        "rest": counts["休息"],
        "weekend_rest": counts["周末休息"],
        "holiday_rest": counts["节假日休息"],
        "annual_leave": counts["年假"],
        "mode": "文字" if use_text else "颜色",
        "daily": daily,
        "has_summary": bool(summary_cols),
    }


def apply_colors(sheet):
    n = 0
    for row in sheet.iter_rows():
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            shift = SHIFT_TEXT_MAP.get(val)
            if shift:
                cell.fill = SHIFT_FILL[shift]
                n += 1
    return n


# ─────────────────────────────────────────────
# API
# ─────────────────────────────────────────────

@app.post("/api/upload")
async def upload(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "只支持 .xlsx 格式")
    data = await file.read()
    session_id = str(uuid.uuid4())
    _file_cache[session_id] = data
    wb = load_workbook(io.BytesIO(data), data_only=True)
    return {"session_id": session_id, "sheets": wb.sheetnames, "filename": file.filename}


@app.get("/api/calc")
async def calc(
    session_id: str,
    sheet: str,
    name: str = "张小燕",
    holidays: str = "",   # 逗号分隔的日期数字，如 "1,2,7"
):
    if session_id not in _file_cache:
        raise HTTPException(404, "session 不存在，请重新上传文件")
    wb = load_workbook(io.BytesIO(_file_cache[session_id]), data_only=True)
    if sheet not in wb.sheetnames:
        raise HTTPException(404, f"Sheet '{sheet}' 不存在")
    holiday_list = [h.strip() for h in holidays.split(",") if h.strip()]
    result = calc_sheet(wb[sheet], name, holiday_list)
    if result is None:
        raise HTTPException(404, f"在 Sheet '{sheet}' 中未找到员工 '{name}'")
    return {"sheet": sheet, "name": name, **result}


@app.post("/api/calc-all")
async def calc_all(
    session_id: str,
    name: str = "张小燕",
    holidays_by_sheet: dict = None,  # {sheet_name: ["1","2",...]} 每个月各自的节假日
):
    if session_id not in _file_cache:
        raise HTTPException(404, "session 不存在，请重新上传文件")
    if holidays_by_sheet is None:
        holidays_by_sheet = {}
    wb = load_workbook(io.BytesIO(_file_cache[session_id]), data_only=True)
    results = []
    grand_actual = 0
    for sname in wb.sheetnames:
        holiday_list = holidays_by_sheet.get(sname, [])
        r = calc_sheet(wb[sname], name, holiday_list)
        if r:
            results.append({"sheet": sname, **r})
            grand_actual += r["actual_work"]
    return {"name": name, "sheets": results, "grand_actual": grand_actual}


@app.get("/api/color-download")
async def color_download(session_id: str, name: str = "张小燕"):
    if session_id not in _file_cache:
        raise HTTPException(404, "session 不存在，请重新上传文件")
    wb = load_workbook(io.BytesIO(_file_cache[session_id]), data_only=False)
    for sname in wb.sheetnames:
        apply_colors(wb[sname])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedule_colored.xlsx"},
    )


@app.get("/api/sheets")
async def get_sheets(session_id: str):
    if session_id not in _file_cache:
        raise HTTPException(404, "session 不存在")
    wb = load_workbook(io.BytesIO(_file_cache[session_id]), data_only=True)
    return {"sheets": wb.sheetnames}
